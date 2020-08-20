# Script de test pour transmettre message de transaction
import datetime, time

from millegrilles import Constantes
from millegrilles.Constantes import SenseursPassifsConstantes
from millegrilles.dao.ConfigurationDocument import ContexteRessourcesDocumentsMilleGrilles
from millegrilles.dao.MessageDAO import BaseCallback
from millegrilles.transaction.GenerateurTransaction import GenerateurTransaction
from threading import Event

from millegrilles.util.JSONEncoders import MongoJSONEncoder

import json

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


contexte = ContexteRessourcesDocumentsMilleGrilles()
contexte.initialiser()


class RequeteMongo(BaseCallback):

    def __init__(self):
        super().__init__(contexte)
        self.contexte.message_dao.register_channel_listener(self)
        self.generateur = GenerateurTransaction(self.contexte)

        self.queue_name = None

        self.channel = None
        self.event_recu = Event()
        self.collection_transactions = self.contexte.document_dao.get_collection(SenseursPassifsConstantes.COLLECTION_TRANSACTIONS_NOM)
        self.collection_documents = self.contexte.document_dao.get_collection(
            SenseursPassifsConstantes.COLLECTION_DOCUMENTS_NOM)

        self.temps_debut_rapport = datetime.datetime(year=2020, month=1, day=1)
        self.temps_fin_rapport = datetime.datetime(year=2020, month=2, day=1)

        self.filtre = {
            'en-tete.domaine': SenseursPassifsConstantes.TRANSACTION_DOMAINE_LECTURE,
            # 'uuid_senseur': {'$in': ['731bf65cf35811e9b135b827eb9064af']},
            SenseursPassifsConstantes.TRANSACTION_DATE_LECTURE: {
                '$gte': self.temps_debut_rapport.timestamp(),
                '$lt': self.temps_fin_rapport.timestamp(),
            },
        }

        self.regroupement_periode = {
            'year': {'$year': '$_evenements._estampille'},
            'month': {'$month': '$_evenements._estampille'},
            'day': {'$dayOfMonth': '$_evenements._estampille'},
            'hour': {'$hour': '$_evenements._estampille'},
        }

        self._regroupement_elem_numeriques = [
            # 'temperature', 'humidite', 'pression', 'millivolt', 'reserve'
            'temperature', 'humidite',
        ]

        # self._accumulateurs = ['max', 'min', 'avg']
        self._accumulateurs = ['avg']

        self.hint = {'_evenements._estampille': -1}

    def on_channel_open(self, channel):
        # Enregistrer la reply-to queue
        self.channel = channel
        channel.queue_declare(durable=True, exclusive=True, callback=self.queue_open_local)

    def queue_open_local(self, queue):
        self.queue_name = queue.method.queue
        print("Queue: %s" % str(self.queue_name))

        self.channel.basic_consume(self.callbackAvecAck, queue=self.queue_name, no_ack=False)
        self.executer()

    def run_ioloop(self):
        self.contexte.message_dao.run_ioloop()

    def deconnecter(self):
        self.contexte.message_dao.deconnecter()

    def traiter_message(self, ch, method, properties, body):
        print("Message recu, correlationId: %s" % properties.correlation_id)
        message = json.loads(str(body, 'utf-8'))
        print(json.dumps(message, indent=4))

    def requete_filtre_1(self):
        curseur_resultat = self.collection_transactions.find(self.filtre).limit(10)
        resultats = list()
        for resultat in curseur_resultat:
            resultats.append({
                'uuid_senseur': resultat['uuid_senseur'],
                'timestamp': resultat['timestamp'],
                'senseurs': resultat['senseurs'],
            })

        print(json.dumps(resultats, indent=2))

    def requete_aggr_1(self):
        regroupement = {
            '_id': {
                'uuid_senseur': '$uuid_senseur',
                'appareil_type': '$senseurs.type',
                'appareil_adresse': '$senseurs.adresse',
                'timestamp': {
                    '$dateFromParts': self.regroupement_periode
                },
            },
        }

        for elem_regroupement in self._regroupement_elem_numeriques:
            for accumulateur in self._accumulateurs:
                key = '%s_%s' % (elem_regroupement, accumulateur)
                regroupement[key] = {'$%s' % accumulateur: '$senseurs.%s' % elem_regroupement}

        operation = [
            {'$match': self.filtre},
            {'$unwind': '$senseurs'},
            {'$group': regroupement},
        ]

        curseur_resultat = self.collection_transactions.aggregate(operation, hint=self.hint)

        resultats = list()
        for resultat in curseur_resultat:
            resultats.append(resultat)

        print(json.dumps(resultats, cls=MongoJSONEncoder, indent=2))
        return resultats

    def projeter_rapport(self):
        resultats = self.requete_aggr_1()

        colonnes = set()
        rangees = dict()
        for resultat in resultats:
            id_result = resultat['_id']
            if id_result.get('appareil_adresse') is not None:
                colonne = id_result['uuid_senseur'] + '/' + id_result['appareil_adresse']
            else:
                colonne = id_result['uuid_senseur'] + '/' + id_result['appareil_type']

            timestamp = id_result['timestamp']
            rangee = rangees.get(timestamp)
            if rangee is None:
                rangee = dict()
                rangees[timestamp] = rangee

            for donnee in resultat.keys():
                if donnee != '_id' and resultat.get(donnee) is not None:
                    colonne_donnee = colonne + '/' + donnee
                    colonnes.add(colonne_donnee)
                    rangee[colonne_donnee] = resultat[donnee]

        colonnes = sorted(colonnes)
        print("Colonnes: " + str(colonnes))

        for timestamp in sorted(rangees.keys()):
            ligne = "Timestamp %s : " % (timestamp)
            rangee = rangees[timestamp]
            for colonne in colonnes:
                if rangee.get(colonne) is not None:
                    ligne = ligne + ', ' + colonne + "=" + str(rangee[colonne])
            print("Timestamp %s : %s" % (timestamp, ligne))

        return rangees, colonnes

    def generer_excel(self):
        rangees, colonnes = self.projeter_rapport()

        wb = Workbook()
        dest_filename = '/home/mathieu/tmp/empty_book.xlsx'
        ws1 = wb.active
        ws1.title = "Pour le fun"

        colonnes = sorted(colonnes)
        senseurs = dict()
        for colonne in colonnes:
            senseur, appareil, mesure = colonne.split('/')
            groupe_appareils = senseurs.get(senseur)
            if groupe_appareils is None:
                groupe_appareils = dict()
                senseurs[senseur] = groupe_appareils
            groupe_mesures = groupe_appareils.get(appareil)
            if groupe_mesures is None:
                groupe_mesures = list()
                groupe_appareils[appareil] = groupe_mesures
            groupe_mesures.append(mesure)

        # Remplacer les ID de senseurs et appareils par leur nom
        filtre = {
            Constantes.DOCUMENT_INFODOC_LIBELLE: SenseursPassifsConstantes.LIBVAL_DOCUMENT_SENSEUR,
            SenseursPassifsConstantes.TRANSACTION_ID_SENSEUR: {'$in': list(senseurs.keys())}
        }
        try:
            curseur_senseurs = self.collection_documents.find(filtre)
            for senseur_db in curseur_senseurs:
                id_senseur = senseur_db[SenseursPassifsConstantes.TRANSACTION_ID_SENSEUR]
                colonnes_senseur = senseurs[id_senseur]
                location_senseur = senseur_db.get(SenseursPassifsConstantes.TRANSACTION_LOCATION)
                if location_senseur is not None:
                    print(location_senseur)
                    # Remplacer le nom du senseur dans la colonne
                    senseurs[location_senseur] = colonnes_senseur
                    del senseurs[id_senseur]
                    for appareil in colonnes_senseur.keys():
                        mesures = colonnes_senseur[appareil]
                        appareil_db = senseur_db['affichage'].get(appareil)
                        if appareil_db is not None:
                            location_appareil = appareil_db.get(SenseursPassifsConstantes.TRANSACTION_LOCATION)
                            if location_appareil is not None:
                                colonnes_senseur[location_appareil] = mesures
                                del colonnes_senseur[appareil]
        except Exception:
            pass
            # Erreur de formattage de l'entete, n'empeche pas de produire le rapport

        # Generer les 3 niveaux d'entete
        no_colonne = 2
        # ws1.cell(column=no_colonne, row=1, value=colonne)
        for senseur in sorted(senseurs.keys()):
            appareils = senseurs[senseur]
            ws1.cell(column=no_colonne, row=1, value=senseur)
            for appareil in sorted(appareils.keys()):
                ws1.cell(column=no_colonne, row=2, value=appareil)
                mesures = appareils[appareil]
                for mesure in mesures:
                    ws1.cell(column=no_colonne, row=3, value=mesure)
                    no_colonne = no_colonne + 1

        ligne = 3
        for timestamp in sorted(rangees.keys()):
            no_colonne = 1
            ligne = ligne + 1
            ws1.cell(column=1, row=ligne, value=timestamp)

            rangee = rangees[timestamp]
            for colonne in colonnes:
                no_colonne = no_colonne + 1
                valeur = rangee.get(colonne)
                if valeur is not None:
                    ws1.cell(column=no_colonne, row=ligne, value=valeur)
                    if 'temperature' in colonne or 'pression' in colonne:
                        ws1.cell(column=no_colonne, row=ligne).number_format = '0.0'
                    else:
                        ws1.cell(column=no_colonne, row=ligne).number_format = '0'

        wb.save(dest_filename)

    def requete_rapport(self):
        fin_rapport = datetime.datetime.utcnow()
        debut_rapport = fin_rapport - datetime.timedelta(days=60)

        transaction = {
            'mesures': ['temperature', 'humidite'],
            'accumulateurs': ['avg', 'max'],
            'senseurs': ['514951f2f43211e99259b827eb53ee51'],
            'groupe_temps': 'hour',
            'periode': {'debut': debut_rapport.timestamp(), 'fin': fin_rapport.timestamp()}
        }

        enveloppe_val = self.generateur.soumettre_transaction(
            transaction, SenseursPassifsConstantes.TRANSACTION_DOMAINE_GENERER_RAPPORT, reply_to=self.queue_name,
            correlation_id='efgh')

        print("Sent: %s" % enveloppe_val)
        return enveloppe_val

    def executer(self):
        self.requete_rapport()


# --- MAIN ---
sample = RequeteMongo()

# TEST
# sample.executer()

# FIN TEST
sample.event_recu.wait(2)
sample.deconnecter()

